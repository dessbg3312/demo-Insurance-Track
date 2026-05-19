#!/usr/bin/env python3
"""
Generic Insurance Schedule Workbook Builder
Usage: python3 build_workbook.py <data.json> <output.xlsx>

Input JSON schema: see references/schema.md
Output: 4-tab Excel workbook (Summary & Totals | Properties | Policies | Auto Insurance)

Requires: openpyxl  (pip install openpyxl --break-system-packages)
"""

import sys
import json
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── colour palette ────────────────────────────────────────────────────
NAVY     = "1F3864"
TEAL     = "2E75B6"
TEAL_LT  = "D6E4F0"
DARK_G   = "375623"
GREEN_LT = "E2EFDA"
AMBER_LT = "FFF9C4"
RED_LT   = "FCE4EC"
GRAY_LT  = "F2F2F2"
WHITE    = "FFFFFF"
PURPLE   = "4A235A"
PURP_LT  = "F3E5F5"
ORANGE   = "C55A11"
ORANGE_LT= "FBE5D6"
GRAY     = "595959"

STATUS_BG = {"Active": GREEN_LT, "Quote": AMBER_LT, "Expired": RED_LT}


# ── style helpers ─────────────────────────────────────────────────────
def _s(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)

def _b(color="BFBFBF"):
    s = _s("thin", color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_b(color="BFBFBF"):
    return Border(left=_s("thin", color), right=_s("thin", color),
                  top=_s("thin", color), bottom=_s("medium", GRAY))

def cs(c, bg=WHITE, fg="000000", bold=False, size=9, wrap=True,
       halign="left", valign="center", b=None, italic=False):
    c.font      = Font(name="Arial", size=size, bold=bold, italic=italic, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    c.border    = b or _b()

def title_row(ws, row, label, ncols, bg=NAVY, size=13, height=26):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    t = ws.cell(row, 1)
    t.value = label
    cs(t, bg=bg, fg=WHITE, bold=True, size=size, halign="center", valign="center",
       wrap=False, b=_b(WHITE))
    ws.row_dimensions[row].height = height

def sub_row(ws, row, label, ncols, bg=TEAL, size=8, height=14):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    s = ws.cell(row, 1)
    s.value = label
    cs(s, bg=bg, fg=WHITE, italic=True, size=size, halign="center", valign="center",
       wrap=False, b=_b(WHITE))
    ws.row_dimensions[row].height = height

def hdr(ws, row, col, label, bg=TEAL, fg=WHITE, size=9, height=None):
    c = ws.cell(row, col)
    c.value = label
    cs(c, bg=bg, fg=fg, bold=True, size=size, halign="center", valign="center",
       b=_b(WHITE))
    if height:
        ws.row_dimensions[row].height = height


def parse_date(val):
    """Accept date object, datetime, or ISO string."""
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    return datetime.strptime(val[:10], "%Y-%m-%d").date()


# ════════════════════════════════════════════════════════════════════
# TAB: PROPERTIES
# ════════════════════════════════════════════════════════════════════
def build_properties(wb, props, portfolio_name):
    ws = wb.active
    ws.title = "Properties"

    COLS = [
        (1,  9,  "Prop ID"),
        (2,  28, "Property Name /\nNickname"),
        (3,  32, "Full Address"),
        (4,  16, "City"),
        (5,  6,  "State"),
        (6,  7,  "ZIP"),
        (7,  18, "Property Type"),
        (8,  8,  "Year\nBuilt"),
        (9,  10, "Units\n(count)"),
        (10, 10, "Sq Ft\n(building)"),
        (11, 28, "Owner / Insured Name"),
        (12, 30, "Mortgagee / Lender"),
        (13, 22, "Primary Agent"),
        (14, 40, "Notes (static)"),
    ]
    NC = len(COLS)
    for ci, w, _ in COLS:
        ws.column_dimensions[get_column_letter(ci)].width = w

    title_row(ws, 1, f"PROPERTIES MASTER — {portfolio_name.upper()} PORTFOLIO", NC, bg=DARK_G)
    sub_row(ws, 2, "Static property data · update only when property characteristics change", NC, bg=DARK_G)
    ws.row_dimensions[3].height = 30
    for ci, _, lbl in COLS:
        hdr(ws, 3, ci, lbl, bg=DARK_G)
    ws.freeze_panes = "A4"

    for i, p in enumerate(props):
        r = i + 4
        ws.row_dimensions[r].height = 30
        bg = WHITE if i % 2 == 0 else GRAY_LT
        vals = [
            p.get("prop_id", ""),
            p.get("nickname", ""),
            p.get("address", ""),
            p.get("city", ""),
            p.get("state", ""),
            p.get("zip", ""),
            p.get("type", ""),
            p.get("year_built"),
            p.get("units"),
            p.get("sqft"),
            p.get("owner", ""),
            p.get("mortgagee", "Not stated in policy"),
            p.get("agent", ""),
            p.get("notes", ""),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci)
            c.value = "Not Found" if val is None else val
            cs(c, bg=bg)


# ════════════════════════════════════════════════════════════════════
# TAB: POLICIES
# ════════════════════════════════════════════════════════════════════
def build_policies(wb, policies, portfolio_name, as_of):
    ws = wb.create_sheet("Policies")

    COLS = [
        (1,  9,  "Prop ID"),
        (2,  20, "Property\nNickname"),
        (3,  26, "Full Address\n(auto from Properties)"),
        (4,  14, "Policy\nNumber"),
        (5,  10, "Status"),
        (6,  26, "Insurance Carrier"),
        (7,  24, "Agency / Broker"),
        (8,  18, "Policy Type"),
        (9,  12, "Effective\nDate"),
        (10, 12, "Expiration\nDate"),
        (11, 12, "Days to\nRenewal"),
        (12, 14, "Total\nPremium ($)"),
        (13, 16, "Building Limit\n(Coverage A)"),
        (14, 20, "Business Income /\nLoss of Rents"),
        (15, 26, "Liability /\nGeneral Aggregate"),
        (16, 18, "Deductible\n(AOP)"),
        (17, 18, "Water Damage\nDeductible"),
        (18, 18, "Sewer Damage\nDeductible"),
        (19, 14, "Units\n(auto)"),
        (20, 12, "Sq Ft\n(auto)"),
        (21, 20, "Price Per Unit\n(auto-calc)"),
        (22, 20, "Price Per Sq Ft\n(auto-calc)"),
        (23, 16, "Inspection\nNoted"),
        (24, 18, "Habitability\nSub-Limit"),
        (25, 40, "Policy PDF\n(Drive link)"),
        (26, 14, "Last\nVerified"),
        (27, 55, "Additional Notes / Endorsements"),
    ]
    NC = len(COLS)
    for ci, w, _ in COLS:
        ws.column_dimensions[get_column_letter(ci)].width = w

    title_row(ws, 1, f"POLICIES — {portfolio_name.upper()} INSURANCE PORTFOLIO", NC)
    sub_row(ws, 2,
            "Dynamic insurance data · update at each renewal · linked to Properties tab via Prop ID · calculated fields auto-update",
            NC)
    ws.row_dimensions[3].height = 32
    for ci, _, lbl in COLS:
        hdr(ws, 3, ci, lbl)
    ws.freeze_panes = "A4"

    # assign group colours by policy number (alternating for ungrouped)
    seen_policies = {}
    palette = ["FFFFFF", "F5F5F5", "E8F5F7", "EAF4E8", "FFF3E0", "E8EAF6"]
    pal_idx = 0
    def group_bg(pol_no):
        nonlocal pal_idx
        if pol_no not in seen_policies:
            seen_policies[pol_no] = palette[pal_idx % len(palette)]
            pal_idx += 1
        return seen_policies[pol_no]

    verified = as_of if isinstance(as_of, date) else parse_date(as_of) or date.today()

    for i, pol in enumerate(policies):
        r = i + 4
        ws.row_dimensions[r].height = 45
        prop_id = pol.get("prop_id")
        pol_no  = pol.get("policy_number", "")
        bg      = group_bg(pol_no)

        # A: Prop ID
        c = ws.cell(r, 1)
        c.value = prop_id if prop_id else "MULTI / N/A"
        cs(c, bg=bg)

        # B: Nickname (vlookup)
        c = ws.cell(r, 2)
        if prop_id:
            c.value = f'=IFERROR(VLOOKUP(A{r},Properties!$A$4:$N$1000,2,FALSE),"—")'
        else:
            c.value = pol.get("nickname_override", "Multiple locations")
        cs(c, bg=bg)

        # C: Address (vlookup)
        c = ws.cell(r, 3)
        if prop_id:
            c.value = f'=IFERROR(VLOOKUP(A{r},Properties!$A$4:$N$1000,3,FALSE),"—")'
        else:
            c.value = pol.get("address_override", "Multiple locations — see Properties tab")
        cs(c, bg=bg)

        # D: Policy number
        c = ws.cell(r, 4)
        c.value = pol_no
        cs(c, bg=bg, bold=True)

        # E: Status
        c = ws.cell(r, 5)
        status = pol.get("status", "Active")
        c.value = status
        cs(c, bg=STATUS_BG.get(status, WHITE), halign="center")

        # F-H: carrier, agency, type
        for ci, key in [(6,"carrier"),(7,"agency"),(8,"policy_type")]:
            ws.cell(r, ci).value = pol.get(key, "")
            cs(ws.cell(r, ci), bg=bg)

        # I: Effective
        c = ws.cell(r, 9)
        c.value = parse_date(pol.get("effective_date"))
        c.number_format = "MM/DD/YYYY"
        cs(c, bg=bg, halign="center")

        # J: Expiration
        c = ws.cell(r, 10)
        c.value = parse_date(pol.get("expiration_date"))
        c.number_format = "MM/DD/YYYY"
        cs(c, bg=bg, halign="center")

        # K: Days to renewal
        c = ws.cell(r, 11)
        c.value = f"=MAX(0,J{r}-TODAY())"
        c.number_format = '0" days"'
        cs(c, bg=bg, halign="center")

        # L: Premium
        c = ws.cell(r, 12)
        c.value = pol.get("premium")
        c.number_format = "$#,##0.00"
        cs(c, bg=bg, halign="right")

        # M: Building limit
        c = ws.cell(r, 13)
        bl = pol.get("building_limit")
        if bl:
            c.value = bl
            c.number_format = "$#,##0"
            cs(c, bg=bg, halign="right")
        else:
            c.value = "N/A"
            cs(c, bg=bg, halign="center")

        # N-R: text fields
        for ci, key in [(14,"business_income"),(15,"liability"),
                        (16,"ded_aop"),(17,"ded_water"),(18,"ded_sewer")]:
            ws.cell(r, ci).value = pol.get(key, "Not Found")
            cs(ws.cell(r, ci), bg=bg)

        # S: Units (vlookup)
        c = ws.cell(r, 19)
        if prop_id:
            c.value = f'=IFERROR(VLOOKUP(A{r},Properties!$A$4:$N$1000,9,FALSE),"—")'
        else:
            c.value = "N/A"
        cs(c, bg=bg, halign="center")

        # T: Sq Ft (vlookup)
        c = ws.cell(r, 20)
        if prop_id:
            c.value = f'=IFERROR(VLOOKUP(A{r},Properties!$A$4:$N$1000,10,FALSE),"—")'
        else:
            c.value = "N/A"
        cs(c, bg=bg, halign="center")

        # U: Price Per Unit
        c = ws.cell(r, 21)
        if prop_id and pol_no not in ("",):
            c.value = f'=IFERROR(IF(S{r}="—","—",ROUND(L{r}/S{r},0)),"—")'
            c.number_format = "$#,##0"
        else:
            c.value = "N/A"
        cs(c, bg=bg, halign="right" if prop_id else "center")

        # V: Price Per Sq Ft
        c = ws.cell(r, 22)
        if prop_id and bl:
            c.value = f'=IFERROR(IF(T{r}="—","—",ROUND(M{r}/T{r},0)),"—")'
            c.number_format = "$#,##0"
        else:
            c.value = "N/A"
        cs(c, bg=bg, halign="right" if (prop_id and bl) else "center")

        # W-X: inspection, habitability
        for ci, key in [(23,"inspection"),(24,"habitability")]:
            ws.cell(r, ci).value = pol.get(key, "Not Found")
            cs(ws.cell(r, ci), bg=bg)

        # Y: PDF link
        pdf = pol.get("pdf_link", "")
        c = ws.cell(r, 25)
        if pdf:
            c.value = "📄 Open Policy Folder"
            c.hyperlink = pdf
            c.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        else:
            c.value = "—"
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _b()

        # Z: Last verified
        c = ws.cell(r, 26)
        c.value = verified
        c.number_format = "MM/DD/YYYY"
        cs(c, bg=bg, halign="center")

        # AA: Notes
        ws.cell(r, 27).value = pol.get("notes", "")
        cs(ws.cell(r, 27), bg=bg)

    # footnote
    note_r = len(policies) + 5
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=NC)
    n = ws.cell(note_r, 1)
    n.value = ("💡  Columns B, C (Address), S (Units), T (Sq Ft), U ($/Unit), V ($/SqFt) are live formulas "
               "linked to the Properties tab — update Properties once and all policy rows update automatically. "
               "Column K (Days to Renewal) always reflects today's date.")
    n.font = Font(name="Arial", size=8, italic=True, color=GRAY)
    n.fill = PatternFill("solid", fgColor="F9F9F9")
    n.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[note_r].height = 28


# ════════════════════════════════════════════════════════════════════
# TAB: AUTO INSURANCE
# ════════════════════════════════════════════════════════════════════
def build_auto(wb, auto_rows, portfolio_name):
    ws = wb.create_sheet("Auto Insurance")

    COLS = [
        (1, 22, "Policy Group /\nInsured"),
        (2, 30, "Named Insured(s)"),
        (3,  9, "State"),
        (4, 30, "Insurance Carrier"),
        (5, 26, "Agency"),
        (6, 22, "Policy Number"),
        (7, 13, "Effective\nDate"),
        (8, 13, "Expiration\nDate"),
        (9, 11, "Days to\nRenewal"),
        (10,14, "Total Premium\n(as stated)"),
        (11,30, "Vehicle(s)"),
        (12,22, "VIN(s)"),
        (13,26, "BI / PD Limits"),
        (14,18, "Comp\nDeductible"),
        (15,18, "Collision\nDeductible"),
        (16,26, "UM/UIM"),
        (17,22, "PIP / Med Pay"),
        (18,55, "Additional Notes"),
    ]
    NC = len(COLS)
    for ci, w, _ in COLS:
        ws.column_dimensions[get_column_letter(ci)].width = w

    title_row(ws, 1, f"AUTO INSURANCE SCHEDULE — {portfolio_name.upper()}", NC)
    sub_row(ws, 2, "Vehicle policies · separate from property insurance · update at each renewal", NC)
    ws.row_dimensions[3].height = 32
    for ci, _, lbl in COLS:
        hdr(ws, 3, ci, lbl)
    ws.freeze_panes = "A4"

    AC = [WHITE, "EDF5FF"]
    for i, row in enumerate(auto_rows):
        r = i + 4
        ws.row_dimensions[r].height = 70
        bg = AC[i % 2]

        vals = [
            row.get("group",""), row.get("insured",""), row.get("state",""),
            row.get("carrier",""), row.get("agency",""), row.get("policy_number",""),
            parse_date(row.get("effective_date")), parse_date(row.get("expiration_date")),
            None,  # days formula
            row.get("premium"), row.get("vehicles",""), row.get("vins",""),
            row.get("bipd",""), row.get("comp_ded",""), row.get("coll_ded",""),
            row.get("um_uim",""), row.get("pip_medpay",""), row.get("notes",""),
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j)
            if j in (7, 8):
                c.value = v
                c.number_format = "MM/DD/YYYY"
                cs(c, bg=bg, halign="center")
            elif j == 9:
                c.value = f"=MAX(0,H{r}-TODAY())"
                c.number_format = '0" days"'
                cs(c, bg=bg, halign="center")
            elif j == 10:
                c.value = v
                c.number_format = "$#,##0.00"
                cs(c, bg=bg, halign="right")
            else:
                c.value = v
                cs(c, bg=bg)


# ════════════════════════════════════════════════════════════════════
# TAB: SUMMARY & TOTALS
# ════════════════════════════════════════════════════════════════════
def build_summary(wb, policies, auto_rows, portfolio_name, as_of_str):
    ws = wb.create_sheet("Summary & Totals", 0)  # first tab

    COL_W = {"A": 30, "B": 22, "C": 10, "D": 30,
             "E": 13, "F": 11, "G": 16, "H": 18, "I": 40}
    for col, w in COL_W.items():
        ws.column_dimensions[col].width = w

    # ── title rows ──────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    t = ws.cell(1, 1)
    t.value = f"SUMMARY & TOTALS — {portfolio_name.upper()} INSURANCE PORTFOLIO"
    cs(t, bg=NAVY, fg=WHITE, bold=True, size=14, halign="center", valign="center",
       wrap=False, b=_b(WHITE))
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:I2")
    s = ws.cell(2, 1)
    s.value = (f"As of {as_of_str}  ·  All premiums as stated in policy documents  ·  "
               "Multi-property policy premiums listed once  ·  Auto premiums as issued")
    cs(s, bg=TEAL, fg=WHITE, italic=True, size=8, halign="center", valign="center",
       wrap=False, b=_b(WHITE))
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6

    # ── KPI card helper ──────────────────────────────────────────────
    def kpi_block(start_col, label, value, sub, bg_lbl, bg_val):
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col+1)
        c = ws.cell(4, start_col)
        c.value = label
        cs(c, bg=bg_lbl, fg=WHITE, bold=True, size=8, halign="center", wrap=False, b=_b(bg_lbl))

        ws.merge_cells(start_row=5, start_column=start_col, end_row=6, end_column=start_col+1)
        v = ws.cell(5, start_col)
        v.value = value
        cs(v, bg=bg_val, bold=True, size=15, halign="center", valign="center",
           wrap=False, b=_b("CCCCCC"))

        ws.merge_cells(start_row=7, start_column=start_col, end_row=7, end_column=start_col+1)
        sv = ws.cell(7, start_col)
        sv.value = sub
        cs(sv, bg=bg_val, fg=GRAY, italic=True, size=7.5, halign="center",
           wrap=False, b=_b("CCCCCC"))

    # classify policies
    def is_multi(pol):
        """Multi-property if the same policy_number appears more than once."""
        return sum(1 for p in policies if p.get("policy_number") == pol.get("policy_number")) > 1

    def is_auto_pol(pol):
        return pol.get("policy_type","").lower() in ("auto","automobile","commercial auto")

    def is_umbrella(pol):
        return "umbrella" in pol.get("policy_type","").lower() or "excess" in pol.get("policy_type","").lower()

    def is_residential(pol):
        t = pol.get("policy_type","").lower()
        return any(x in t for x in ["ho-3","ho-6","homeowner","condo","cea","earthquake","rental"])

    def is_commercial(pol):
        return not is_umbrella(pol) and not is_residential(pol) and not is_auto_pol(pol)

    # build unique-policy premium totals (avoid double-counting multi-property)
    seen = set()
    comm_bound = res_bound = umb_bound = quote_total = 0
    comm_props = res_props = quote_props = 0

    for pol in policies:
        pno = pol.get("policy_number","")
        prem = pol.get("premium") or 0
        status = pol.get("status","Active")
        if pno in seen:
            continue
        seen.add(pno)
        if status == "Quote":
            quote_total += prem
            quote_props += 1
        elif is_commercial(pol):
            comm_bound += prem
            comm_props += 1
        elif is_umbrella(pol):
            umb_bound += prem
        elif is_residential(pol):
            res_bound += prem
            res_props += 1

    auto_total = sum((a.get("premium") or 0) for a in auto_rows)
    auto_veh   = sum(
        len([v for v in a.get("vehicles","").split(";") if v.strip()])
        for a in auto_rows
    )

    def fmt_k(v):
        return f"${v:,.0f}"

    kpi_block(1, "COMMERCIAL (BOUND)",   fmt_k(comm_bound + umb_bound),
              f"{comm_props} policies", TEAL, TEAL_LT)
    kpi_block(3, "RESIDENTIAL",          fmt_k(res_bound),
              f"{res_props} policies", DARK_G, GREEN_LT)
    kpi_block(5, "AUTO (AS ISSUED)",     fmt_k(auto_total),
              f"{len(auto_rows)} policies · {auto_veh} vehicles", ORANGE, ORANGE_LT)
    kpi_block(7, "QUOTE / NOT BOUND",    fmt_k(quote_total),
              f"{quote_props} quotes", PURPLE, PURP_LT)

    # fill col I on KPI rows
    for rr in [4, 5, 6, 7]:
        c = ws.cell(rr, 9)
        cs(c, bg=GRAY_LT, b=_b("E0E0E0"))

    for rr in [4,5,6,7]:
        ws.row_dimensions[rr].height = {4:16,5:22,6:22,7:14}[rr]
    ws.row_dimensions[8].height = 6

    # ── column headers row 9 ─────────────────────────────────────────
    HDRS = ["Category / Policy Group","Policy Number","Status",
            "Properties / Locations","Expiration","Days Left",
            "Premium (as stated)","Building Limit (Cov A)","Notes"]
    ws.row_dimensions[9].height = 28
    for ci, lbl in enumerate(HDRS, 1):
        hdr(ws, 9, ci, lbl, bg=NAVY)

    ws.freeze_panes = "A10"

    ROW = 10

    def section_hdr(label, bg=TEAL, fg=WHITE):
        nonlocal ROW
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=9)
        c = ws.cell(ROW, 1)
        c.value = "  " + label
        cs(c, bg=bg, fg=fg, bold=True, size=9, halign="left",
           valign="center", b=_b(fg))
        ws.row_dimensions[ROW].height = 17
        ROW += 1

    def data_row(cat, pol_no, status, props_lbl, exp_date, prem, bldg, notes, bg=WHITE):
        nonlocal ROW
        ws.row_dimensions[ROW].height = 38
        ws.cell(ROW, 1).value = cat;       cs(ws.cell(ROW,1), bg=bg)
        ws.cell(ROW, 2).value = pol_no;    cs(ws.cell(ROW,2), bg=bg, bold=True)
        c3 = ws.cell(ROW, 3)
        c3.value = status
        cs(c3, bg=STATUS_BG.get(status, WHITE), halign="center")
        ws.cell(ROW, 4).value = props_lbl; cs(ws.cell(ROW,4), bg=bg)
        c5 = ws.cell(ROW, 5)
        c5.value = exp_date
        c5.number_format = "MM/DD/YYYY"
        cs(c5, bg=bg, halign="center")
        c6 = ws.cell(ROW, 6)
        c6.value = f"=MAX(0,E{ROW}-TODAY())"
        c6.number_format = '0" days"'
        cs(c6, bg=bg, halign="center")
        c7 = ws.cell(ROW, 7)
        c7.value = prem
        c7.number_format = "$#,##0.00"
        cs(c7, bg=bg, halign="right")
        c8 = ws.cell(ROW, 8)
        if bldg:
            c8.value = bldg
            c8.number_format = "$#,##0"
            cs(c8, bg=bg, halign="right")
        else:
            c8.value = "N/A"
            cs(c8, bg=bg, halign="center")
        ws.cell(ROW, 9).value = notes;     cs(ws.cell(ROW,9), bg=bg)
        ROW += 1

    def subtotal_row(label, prem, bldg, bg=TEAL_LT):
        nonlocal ROW
        ws.row_dimensions[ROW].height = 16
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
        c = ws.cell(ROW, 1)
        c.value = label
        cs(c, bg=bg, bold=True, size=9, b=_thick_b())
        c7 = ws.cell(ROW, 7)
        c7.value = prem
        c7.number_format = "$#,##0.00"
        cs(c7, bg=bg, bold=True, halign="right", b=_thick_b())
        c8 = ws.cell(ROW, 8)
        if bldg:
            c8.value = bldg
            c8.number_format = "$#,##0"
            cs(c8, bg=bg, bold=True, halign="right", b=_thick_b())
        else:
            c8.value = "—"
            cs(c8, bg=bg, halign="center", b=_thick_b())
        cs(ws.cell(ROW, 9), bg=bg, b=_thick_b())
        ROW += 1

    def grand_row(label, prem, bldg, fg=WHITE, size=11):
        nonlocal ROW
        ws.row_dimensions[ROW].height = 22
        ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
        c = ws.cell(ROW, 1)
        c.value = "  " + label
        cs(c, bg=NAVY, fg=fg, bold=True, size=size, b=_b(WHITE))
        c7 = ws.cell(ROW, 7)
        c7.value = prem
        c7.number_format = "$#,##0.00"
        cs(c7, bg=NAVY, fg=fg, bold=True, size=size, halign="right", b=_b(WHITE))
        c8 = ws.cell(ROW, 8)
        if bldg:
            c8.value = bldg
            c8.number_format = "$#,##0"
            cs(c8, bg=NAVY, fg=fg, bold=True, size=size, halign="right", b=_b(WHITE))
        else:
            c8.value = "—"
            cs(c8, bg=NAVY, fg=fg, size=size, halign="center", b=_b(WHITE))
        cs(ws.cell(ROW, 9), bg=NAVY, fg=fg, b=_b(WHITE))
        ROW += 1
        # spacer
        ws.row_dimensions[ROW].height = 5
        ROW += 1

    # ── build sections from unique policies ───────────────────────────
    processed = set()

    def emit_section(pols, section_label, bg_hdr=TEAL):
        if not pols:
            return
        section_hdr(section_label, bg=bg_hdr)
        sub_prem = 0
        sub_bldg = 0
        alt = [WHITE, GRAY_LT]
        for idx, pol in enumerate(pols):
            pno = pol.get("policy_number","")
            if pno in processed:
                continue
            processed.add(pno)
            prem = pol.get("premium") or 0
            bldg = pol.get("building_limit")
            # aggregate bldg limits for same policy across multiple rows
            same = [p for p in policies if p.get("policy_number") == pno]
            agg_bldg = sum(p.get("building_limit") or 0 for p in same) or None
            prop_ids = ", ".join(filter(None, [p.get("prop_id") for p in same]))
            exp = parse_date(pol.get("expiration_date"))
            notes = pol.get("notes","")
            data_row(
                cat=pol.get("carrier","") + " — " + ", ".join(
                    p.get("nickname","") or p.get("address","") for p in same[:3]),
                pol_no=pno,
                status=pol.get("status","Active"),
                props_lbl=prop_ids if prop_ids else "MULTI / N/A",
                exp_date=exp,
                prem=prem,
                bldg=agg_bldg,
                notes=notes,
                bg=alt[idx % 2],
            )
            sub_prem += prem
            sub_bldg += agg_bldg or 0
        subtotal_row(f"SUBTOTAL — {section_label}", sub_prem, sub_bldg or None)
        return sub_prem, sub_bldg or None

    # separate into buckets
    multi_pols  = []
    single_pols = []
    umb_pols    = []
    res_pols    = []
    quote_pols  = []
    seen_pno    = set()

    for pol in policies:
        pno = pol.get("policy_number","")
        if pno in seen_pno:
            continue
        seen_pno.add(pno)
        status = pol.get("status","Active")
        if status == "Quote":
            quote_pols.append(pol)
        elif is_umbrella(pol):
            umb_pols.append(pol)
        elif is_residential(pol):
            res_pols.append(pol)
        elif is_commercial(pol):
            if is_multi(pol):
                multi_pols.append(pol)
            else:
                single_pols.append(pol)

    comm_sub_prem = 0
    comm_sub_bldg = 0

    if multi_pols:
        r = emit_section(multi_pols, "COMMERCIAL — MULTI-PROPERTY POLICIES")
        if r:
            comm_sub_prem += r[0]; comm_sub_bldg += r[1] or 0

    if single_pols:
        r = emit_section(single_pols, "COMMERCIAL — SINGLE PROPERTY")
        if r:
            comm_sub_prem += r[0]; comm_sub_bldg += r[1] or 0

    if umb_pols:
        r = emit_section(umb_pols, "LIABILITY UMBRELLA")
        if r:
            comm_sub_prem += r[0]

    if comm_sub_prem:
        grand_row("GRAND TOTAL — COMMERCIAL (BOUND)",
                  comm_sub_prem, comm_sub_bldg or None)

    if quote_pols:
        emit_section(quote_pols, "QUOTE / NOT YET BOUND", bg_hdr=PURPLE)

    if res_pols:
        emit_section(res_pols, "RESIDENTIAL / PERSONAL", bg_hdr=DARK_G)

    # auto section
    if auto_rows:
        section_hdr("AUTO INSURANCE  (see Auto Insurance tab for full detail)", bg=ORANGE)
        auto_sub = 0
        for idx, a in enumerate(auto_rows):
            exp = parse_date(a.get("expiration_date"))
            prem = a.get("premium") or 0
            data_row(
                cat=a.get("carrier","") + " — " + a.get("group",""),
                pol_no=a.get("policy_number",""),
                status="Active",
                props_lbl=a.get("vehicles",""),
                exp_date=exp,
                prem=prem,
                bldg=None,
                notes=a.get("notes",""),
                bg=[WHITE, GRAY_LT][idx % 2],
            )
            auto_sub += prem
        subtotal_row("SUBTOTAL — AUTO (as issued; 6-mo policies not annualized)",
                     auto_sub, None, bg="FBE5D6")

    all_bound = comm_sub_prem + res_bound + auto_total
    grand_row("★  PORTFOLIO GRAND TOTAL (ALL BOUND, EXCL. QUOTE)",
              all_bound, comm_sub_bldg + sum(
                  (p.get("building_limit") or 0) for p in res_pols) or None)

    # footnote
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=9)
    fn = ws.cell(ROW, 1)
    fn.value = (
        "NOTES: (1) Multi-property premiums listed once per policy, not per row. "
        "(2) 6-month auto premiums shown as issued, not annualized. "
        "(3) Building Limits = Coverage A only. "
        "(4) ⚠ = item requiring follow-up. "
        "(5) Days to Renewal is always live (calculated from today's date)."
    )
    fn.font = Font(name="Arial", size=7.5, italic=True, color=GRAY)
    fn.fill = PatternFill("solid", fgColor=GRAY_LT)
    fn.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[ROW].height = 28


# ════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════
def main():
    if len(sys.argv) < 3:
        print("Usage: python3 build_workbook.py <data.json> <output.xlsx>")
        sys.exit(1)

    data_path, out_path = sys.argv[1], sys.argv[2]

    with open(data_path) as f:
        data = json.load(f)

    portfolio  = data.get("portfolio_name", "Portfolio")
    as_of      = data.get("as_of_date", str(date.today()))
    props      = data.get("properties", [])
    policies   = data.get("policies", [])
    auto_rows  = data.get("auto_policies", [])

    wb = Workbook()

    # Build tabs in this order, Summary inserts itself at position 0
    build_properties(wb, props, portfolio)
    build_policies(wb, policies, portfolio, as_of)
    build_auto(wb, auto_rows, portfolio)
    build_summary(wb, policies, auto_rows, portfolio, as_of)

    # Enforce tab order
    desired = ["Summary & Totals", "Properties", "Policies", "Auto Insurance"]
    wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 99)

    wb.save(out_path)
    print(f"✓ Saved: {out_path}")
    print(f"  Properties: {len(props)} rows")
    print(f"  Policies:   {len(policies)} rows")
    print(f"  Auto:       {len(auto_rows)} rows")


if __name__ == "__main__":
    main()
